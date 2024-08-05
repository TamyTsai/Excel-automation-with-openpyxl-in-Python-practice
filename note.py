from openpyxl import Workbook, load_workbook
# Workbook為Excel檔案之意
# openpyxl只支援2010以上的excel版本

# 將數字轉為欄位英文字母的工具(1=A, 2=B ...)
from openpyxl.utils import get_column_letter

# 檔案

# C 創建Excel檔案
wb = Workbook()

# R 讀取Excel檔案
wb = load_workbook('profile.xlsx')
wb = load_workbook('new.xlsx')


# 工作表

# C 創建工作表(但還沒儲存，儲存完才會真的修改)
wb.create_sheet('第四個工作表')

# R 讀取預設工作表
ws = wb.active
# R 讀取特定工作表
ws = wb['工作表2']
# print(ws)
# <Worksheet "工作表2">
# R 讀取檔案中 所有工作表(列表)
print(wb.sheetnames)

# U 更改工作表名稱
ws.title = 'first'

# 資料

# R 讀取工作表中的某一格的值
print(ws['A5'].value)

# C、U 修改、填入 工作表中的某一格的值(但還沒儲存，儲存完才會真的修改)
ws['A5'].value = '小灰'
ws['A1'].value = 100

# C 一次新增多列多欄資料
ws.append([123,456,789,0])
ws.append([168,78423,721])
ws.append([26,457])

# R 讀取工作表中的某一範圍的值
for row in range(1,4): # 3列(不含最後一個數)
    for col in range(1,5): # 4欄
        char = get_column_letter(col) # 會回傳每欄對應的英文字母(第1欄為A 第2欄為B...)
        print(ws[char + str(row)].value)

# U 一次修改某一範圍的資料
for row in range(1,4): # 3列(不含最後一個數)
    for col in range(1,5): # 4欄
        char = get_column_letter(col) # 會回傳每欄對應的英文字母(第1欄為A 第2欄為B...)
        ws[char + str(row)].value = char + str(row)

# U 合併儲存格
ws.merge_cells('A1:D2')
# U 回復已經合併之儲存格
ws.unmerge_cells('A1:D1')

# U 移動資料
ws.move_range('A3:D4', rows=2, cols=3) # 橫排往下移兩格，直排往右移三格(負數代表往上或往左)

# U 插入橫排
ws.insert_rows(3) # 在第3列插入橫排
# U 插入直排
ws.insert_cols(2) # 2為B
# D 刪除橫排
ws.delete_rows(3)
# D 刪除直排
ws.delete_cols(2)


# 儲存檔案(不能在檔案開啟時運行 會發生衝突)(C、R、U都要儲存才會生效)
wb.save('profile.xlsx')
wb.save('new.xlsx') # 原本就存在同名檔案的話，會覆蓋同名檔案




