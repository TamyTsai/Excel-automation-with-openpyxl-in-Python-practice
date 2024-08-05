from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
# 引入字體物件
from openpyxl.styles import Font

# 創建Excel檔案
wb = Workbook()

# 讀取預設工作表
ws = wb.active

# 一次新增一列多欄資料
title = ['姓名', '身高', '年紀', '體重']
ws.append(title)

# 一次新增多列多欄資料
data = [ # 裝著字典的列表
   {
        'name': '小白',
        'tall': 180,
        'age': 23,
        'weight': 74
    },
    {
        'name': '小黃',
        'tall': 177,
        'age': 28,
        'weight': 90
    },
    {
        'name': '小綠',
        'tall': 160,
        'age': 30,
        'weight': 60
    },
    {
        'name': '小灰',
        'tall': 155,
        'age': 50,
        'weight': 50
    },
    {
        'name': '小黑',
        'tall': 170,
        'age': 46,
        'weight': 99
    }
]

for person in data: # data是裝有字典的陣列 每個person代表一個字典
    ws.append(list(person.values())) 
    # 將字典中的每個value取出來 然後轉換成列表(如:['小白',180,23,74])

for col in range(2, 5):
    char = get_column_letter(col)
    # 在第2~4欄的第7列
    ws[char + '7'] = f'=AVERAGE({char + "2"}:{char + "6"})' # f搭配花括弧可以讓變數與自串串接

# 將第一列字體變成粗體
for col in range(1,5):
    char = get_column_letter(col)
    # 更改 第1~4欄的第1列 之字體、顏色(working with styles)(https://openpyxl.readthedocs.io/en/stable/styles.html)
    ws[char + '1'].font = Font(bold=True, color="00000080")

wb.save('practice_data.xlsx')